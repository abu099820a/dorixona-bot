"""
exclusive_handlers.py — Eksklyuziv dori: filiallar bo'yicha "yaxshi ketmoqda /
ketmagan" tahlili va yaqin (dori hali yo'q) aptekalar ro'yxatini chiqarish.

Ishlash tartibi (admin uchun):
1. Admin botga /eksklyuziv (kerak bo'lsa radius bilan: /eksklyuziv 1.5) yozadi.
2. Bot shu dorining "aylanma" (turnover) hisobotini .xlsx fayl sifatida
   yuborishni so'raydi (1C/boshqa tizimdan olingan, "Оборот по сети" formatidagi
   hisobot — har bir filial uchun: qoldiq boshi, sotib olindi, SOTUV, ko'chirildi,
   qoldiq oxiri ustunlari bilan).
3. Admin faylni yuboradi.
4. Bot:
   - Filiallarni SOTILGAN miqdoriga qarab toifalarga ajratadi
     (Yaxshi ketmoqda / Sekin ketmoqda / Ketmagan / Yangi berilgan).
   - Filial nomlarini joylashuv ma'lumotlari bilan (Sheet1 — bot.py dagi
     load_df() ishlatadigan bir xil manba, Google Sheets export orqali,
     gspread credential SHART EMAS) mos keladi.
   - "Yaxshi ketmoqda" filiallarga berilgan radius (km) ichida joylashgan,
     bu dori HALI YO'Q aptekalarni topadi.
   - Natijani xabar + tayyor .xlsx fayl sifatida qaytaradi.

MUHIM: bu modul PHARMACY_SHEET_ID/SHEETS_ID dagi "Sheet1" varag'ini FAQAT
o'qish uchun ishlatadi (yozish YO'Q) — shuning uchun registratsiya/davomat
modullaridagi qaysi varaq nima uchun ishlatilishi bilan hech qanday
to'qnashuv yo'q.
"""

import os
import io
import re
import math
import logging

import requests
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from thefuzz import fuzz, process as fuzz_process

from telegram import Update
from telegram.ext import MessageHandler, CommandHandler, ContextTypes, filters

from attendance import run_read
from salary_handlers import ADMIN_IDS

logger = logging.getLogger(__name__)

SHEETS_ID = os.getenv("SHEETS_ID", "1CfuogH-yY--y5kiBK0qXsl5AFi_Hmzj_onWcA-Qyvco")
LOCATION_SHEET_NAME = "Sheet1"
DEFAULT_RADIUS_KM = 2.0

# Suhbat holati (bot.py dagi boshqa state raqamlari bilan TO'QNASHMASLIGI
# uchun 700 dan boshlangan; agar bot.py da allaqachon 700+ band bo'lsa,
# shu raqamni bot.py bilan birga o'zgartiring).
EKS_WAIT_XLSX = 700

# Ba'zi filiallar nomi aylanma hisobotida va Sheet1'da tizimli ravishda
# har xil yozilishi ma'lum bo'lgan holatlar uchun qo'lda mos yozuv.
# thefuzz bularni ishonchli aniqlay olmaydi (masalan bir nechta "гор
# больница" nuqtasi bo'lgani uchun).
#
# MUHIM (2026-08 tuzatish): bu lug'atning ESKI versiyasida barcha 6 ta
# yozuv TESKARI yo'nalishda edi — kalit "haqiqiy" (Sheet1'dagi) nomga
# o'xshab yozilgan, qiymat esa aylanma hisobotdagi nomga o'xshab yozilgan
# edi. Lekin _match_location() kodi buni ALIAS[filial_raw] -> Sheet1'dagi
# nom deb ishlatadi (pastga qarang) — ya'ni yo'nalish TAM TESKARI bo'lgani
# uchun bu 6 ta alias amalda HECH QACHON ishlamagan (aylanma hisobotdagi
# asl matn hech qachon lug'at kalitiga to'g'ri kelmagan). Mijoz yuborgan
# ikkita haqiqiy ro'yxatni (aylanma hisobotdagi nomlar va Sheet1'dagi
# haqiqiy nomlar) solishtirib, TO'G'RI yo'nalishda qayta tuzildi va yana
# ko'plab yangi (avval umuman aliassiz bo'lgan) nomlar ham qo'shildi.
#
# Format: "aylanma hisobotdagi xom nom (kichik harf)" -> "Sheet1'dagi
# ANIQ nom (katta harf, name_to_row kalitlari bilan bir xil formatda)".
KNOWN_ALIASES = {
    "(1) гор больница": "1 ГОР БОЛЬНИЦА (1)",
    "(2) гор больница": "1 ГОР БОЛЬНИЦА (2)",
    "фарм люкс": "1 ГОР БОЛЬНИЦА ФАРМ ЛЮКС",
    "ташселмаш": "ТАШСЕЛЬМАШ",
    "юнусабад ахмад дониш": "АХМАД ДОНИШ",
    "кадешва бозор": "КАДЫШЕВА БАЗАР",
    "янги хаёт элит маркет": "ЯНГИ ХАЕТ ЭЛИТ МАРКЕТ",
    "алфраганус": "АЛЬФРАГАНУС",
    "чиланзар ал-хоразмий": "ЧИЛАНЗАР АЛЬ ХОРАЗМИ",
    "шахристанский": "ШАХРИСТАНСКАЯ",
    "улугбек пасёлка": "ПОСЕЛОК УЛУГБЕК",
    "самарканд фрунзиский": "САМАРКАНД ФРУНЗЕНСКАЯ",
    "октепа чилонзор": "ЧИЛАНЗАР ОКТЕПА",
    "янги хаёт 9-худуд": "ЯНГИ ХАЕТ 9 ХУДУД",
    "литературный учтепа": "УЧТЕПА ЛИТЕРАТУРНАЯ",
    "ширин чилонзор": "ЧИЛАНЗАР ШИРИН",
    "альгоритм гулистон": "АЛГОРИТМ ГУЛИСТОН",
    "корасу-3": "КОРАСУВ-3",
    "корасу-2 садаф": "КОРАСУВ-САДАФ",
    "м.улугбек налоговый": "М.УЛУГБЕК НАЛОГОВАЯ",
    "кора камиш тансикбоев": "КАРА КАМЫШ ТАНСИКБОЕВ",
    "янги хаёт спутник 7-дахаси": "ЯНГИХАЁТ СПУТНИК 7-ДАХАСИ",
    "куйи чирчиқ 5-лет": "КУЙИ ЧИРЧИК 5-ЛЕТ",
    "наманган тўракўргон болница": "НАМАНГАН ТУРАКУРГОН БОЛНИЦА",
    "ўрикзор бозор": "УРИКЗОР БОЗОР",
    "чинабад юнусобод": "ЧИНОБОД ЮНУСАБАД",
    "шахрисабз гор больница": "КАШКАДАРЁ ШАХРИСАБЗ ГОР БОЛЬНИЦА",
    "сергели-7 бозор": "СЕРГЕЛИ 7 БАЗАР",
    "юнусабад 18 квартал": "ЮНУСАБАД 18",
    "чиланзар катта кани": "КАТТА КАНИ",
    "мега планет юнусобод": "МЕГА ПЛАНЕТ",
    "юнусабад 4- квартал": "ЮНУСАБАД 4",
    "наманган қала": "НАМАНГАН КАЛА",
    "косон бозор": "КАШКАДАРЁ КОСОН БОЗОР",
    "қўқон тез ёрдам": "КУКОН ТЕЗ ЁРДАМ",
    "нишон": "КАШКАДАРЁ НИШОН",
    # Foydalanuvchi tomonidan aniqlashtirilgan (bir nechta ehtimoliy
    # filial bo'lgani uchun avtomatik aniqlab bo'lmasdi):
    "термез": "ТЕРМИЗ АЙРИТОМ",
    "сергели-7": "СЕРГЕЛИ 7 БАЗАР",
}

# MUHIM: quyidagi nomlar aylanma hisobotda uchraydi, lekin Sheet1'dagi
# hech bir filial nomiga (hatto taxminan ham) mos kelmadi — ya'ni bu
# filiallar Google Sheet'dagi lokatsiya jadvalida UMUMAN topilmadi.
# Ular hali ham "unmatched" (lokatsiyasi topilmagan) sifatida chiqadi:
#   - СЕРГЕЛИ -1 БОЗОРЧА
#   - ҚУЙЛИҚ МАССИВ-5
#   - ОЛМАЗОР ЖК БОУРУМ
#   - ГАНГА
# Bularni Sheet1'da qidirib, agar mavjud bo'lsa aniq nomini toping va
# yuqoridagi KNOWN_ALIASES ro'yxatiga qo'shib bering — yoki bu filiallar
# hali Sheet1'ga umuman kiritilmagan bo'lsa, avval o'sha yerga (lokatsiya
# jadvaliga) qo'shish kerak bo'ladi.
FUZZY_MATCH_THRESHOLD = 80  # shundan past bo'lsa "tekshirish kerak" deb belgilanadi
SKIP_FILIAL_NAMES = {"по сети", "склад"}


def haversine_km(lat1, lon1, lat2, lon2):
    R = 6371.0
    p1, p2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi / 2) ** 2 + math.cos(p1) * math.cos(p2) * math.sin(dlmb / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def _load_location_df():
    """
    Filiallar joylashuvini (Filial №, Nomi, Latitude, Longitude) o'qiydi.
    bot.py dagi load_df() bilan AYNAN bir xil manba/usul: Google Sheets
    export (HTTP), gspread credential shart emas. Bu funksiya BLOKLOVCHI
    (tarmoq so'rovi) — faqat run_read() orqali chaqirilishi kerak.
    """
    url = f"https://docs.google.com/spreadsheets/d/{SHEETS_ID}/export?format=xlsx&sheet={LOCATION_SHEET_NAME}"
    resp = requests.get(url, timeout=30)
    resp.raise_for_status()
    df = pd.read_excel(io.BytesIO(resp.content)).fillna("")
    required = ["Filial №", "Nomi (UZ)", "Latitude", "Longitude"]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise RuntimeError(f"Sheet1 da kerakli ustunlar topilmadi: {missing}")
    return df


def classify(sold: int, ost_boshi, peremeshenie) -> str:
    sold = abs(sold or 0)
    ost_boshi = ost_boshi or 0
    peremeshenie = peremeshenie or 0
    if sold >= 3:
        return "Yaxshi ketmoqda"
    if 1 <= sold <= 2:
        return "Sekin ketmoqda"
    if sold == 0 and ost_boshi > 0:
        return "Ketmagan (zaxira bor, sotuv yo'q)"
    if sold == 0 and ost_boshi == 0 and peremeshenie > 0:
        return "Yangi berilgan (hali erta baholash uchun)"
    return "Aniqlanmagan"


def _parse_turnover_workbook(xlsx_bytes: bytes):
    """
    MUHIM (2026-08, ko'p dorili fayllar uchun qo'shildi): bitta firmaning
    bir nechta dorisi bo'lsa, aylanma hisobot BITTA xlsx faylda, lekin har
    bir dori ALOHIDA varaqda (sheet) keladi (mijoz tasdiqlagan format).
    Avval kod faqat wb.worksheets[0] (birinchi varaq)ni o'qirdi — agar
    faylda 2+ varaq bo'lsa, qolgan dorilar UMUMAN e'tiborga olinmasdi.

    Endi BARCHA varaqlar aylanib chiqiladi, har biri alohida dori sifatida
    parse qilinadi. Ma'lumotsiz (bo'sh, sarlavha-only) varaqlar avtomatik
    o'tkazib yuboriladi. Qaytaradi: [(product_name, data_rows), ...] —
    ro'yxat, har bir element bitta varaq/doriga tegishli.
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    results = []
    for ws in wb.worksheets:
        product_name, data_rows = _parse_turnover_sheet(ws)
        if data_rows:
            results.append((product_name, data_rows))
    return results


def _parse_turnover_sheet(ws):
    """Bitta varaqdan (bir doriga tegishli) filiallar bo'yicha qatorlarni o'qiydi."""
    rows = list(ws.iter_rows(min_row=1, values_only=True))
    # Birinchi ustundagi qiymat dori nomi bo'lgan qatorni topamiz (2-3 qatorlarda
    # sarlavha bo'lishi mumkin, shuning uchun filial nomi bo'sh bo'lmagan va
    # "по сети"/"склад" bo'lmagan birinchi qatordan boshlaymiz).
    product_name = None
    data_rows = []
    for r in rows:
        if r is None or len(r) < 12:
            continue
        filial = r[2]
        if not filial:
            continue
        ost_boshi, sotuv, peremeshenie, ost_oxiri = r[3] or 0, r[7] or 0, r[9] or 0, r[11] or 0
        # Faqat RAQAMLI qiymatga ega qatorlar haqiqiy ma'lumot qatori
        # hisoblanadi — shu tekshiruv orqali sarlavha qatorlarini (ustun
        # nomlari matn bo'lgani uchun) chetlab o'tamiz. MUHIM: mahsulot
        # nomini ('product_name') ham FAQAT shu tasdiqlangan qatordan
        # olamiz — aks holda 1-qatordagi ustun sarlavhasi ("наименование")
        # xato ravishda mahsulot nomi sifatida qabul qilinib qolardi.
        if not isinstance(ost_boshi, (int, float)):
            continue
        if product_name is None and r[0]:
            product_name = r[0]
        if filial in SKIP_FILIAL_NAMES:
            continue
        data_rows.append({
            "filial_raw": str(filial).strip(),
            "ost_boshi": ost_boshi, "sotuv": sotuv,
            "peremeshenie": peremeshenie, "ost_oxiri": ost_oxiri,
        })
    return product_name or "Nomalum dori", data_rows


def _norm_name(text) -> str:
    """
    Sheet1'dagi "Nomi (UZ)" ustunida ba'zi qatorlarda ORTIQCHA (ikkitadan
    ortiq) bo'shliq uchrashi aniqlandi (masalan "ЧИНОБОД  ЮНУСАБАД" —
    so'zlar orasida bitta emas, ikkita bo'shliq). Bunday holatda oddiy
    .strip().upper() YETARLI EMAS — chunki KNOWN_ALIASES ichidagi target
    (bitta bo'shliq bilan yozilgan) name_to_row lug'atidagi kalitga
    (ikkita bo'shliq bilan) mos kelmay, alias "topilmadi" deb chiqib
    ketardi. Shu funksiya barcha ustunlar ORASIDAGI bo'shliqlarni ham
    bittaga tenglashtiradi — shunda ikkala tomon (Sheet1'dan o'qilgan nom
    va KNOWN_ALIASES'dagi target) doim bir xil formatga keladi.
    """
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip()).upper()


def _match_location(filial_raw: str, loc_df, name_choices, name_to_row):
    key = filial_raw.strip().lower()
    if key in KNOWN_ALIASES:
        target = _norm_name(KNOWN_ALIASES[key])
        row = name_to_row.get(target)
        if row is not None:
            return row, "alias", 100
    best = fuzz_process.extractOne(filial_raw, name_choices, scorer=fuzz.token_sort_ratio)
    if best is None:
        return None, "no_match", 0
    name, score = best[0], best[1]
    row = name_to_row.get(_norm_name(name))
    how = "fuzzy" if score >= FUZZY_MATCH_THRESHOLD else "low_confidence"
    return row, how, score


def _analyze_product(turnover_rows, loc_df, name_choices, name_to_row, radius_km):
    """
    Bitta doriga tegishli aylanma qatorlarini (turnover_rows) tahlil qiladi:
    lokatsiya bilan moslashtiradi, toifalarga ajratadi, va shu DORIGA
    tegishli "yaqin nomzod aptekalar"ni topadi. MUHIM: nomzodlarni
    tekshirishda faqat SHU dori bor filiallar (has_drug_nos) chetlab
    o'tiladi — boshqa dori bor-yo'qligi bu yerda ahamiyatsiz, chunki har
    bir dori uchun "hali yo'q" degani BOSHQA-BOSHQA bo'lishi mumkin.
    """
    matched, unmatched = [], []
    for t in turnover_rows:
        row, how, score = _match_location(t["filial_raw"], loc_df, name_choices, name_to_row)
        if row is None or how == "low_confidence" or how == "no_match":
            t["best_guess"] = row["Nomi (UZ)"] if row is not None else ""
            t["match_score"] = score
            unmatched.append(t)
            continue
        t["category"] = classify(t["sotuv"], t["ost_boshi"], t["peremeshenie"])
        t["loc_nom"] = row["Nomi (UZ)"]
        t["loc_no"] = row["Filial №"]
        t["lat"] = float(row["Latitude"])
        t["lon"] = float(row["Longitude"])
        matched.append(t)

    has_drug_nos = set(t["loc_no"] for t in matched)
    good = [t for t in matched if t["category"] == "Yaxshi ketmoqda"]

    candidates_by_no = {}
    for t in good:
        for _, cd in loc_df.iterrows():
            cd_no = cd["Filial №"]
            if cd_no in has_drug_nos:
                continue
            try:
                clat, clon = float(cd["Latitude"]), float(cd["Longitude"])
            except (TypeError, ValueError):
                continue
            dist = haversine_km(t["lat"], t["lon"], clat, clon)
            if dist <= radius_km:
                key = cd_no
                cand = {
                    "candidate_nom": cd["Nomi (UZ)"], "candidate_no": cd_no,
                    "candidate_lat": clat, "candidate_lon": clon,
                    "good_filial": t["filial_raw"], "good_sold": abs(t["sotuv"]),
                    "distance_km": round(dist, 2),
                }
                if key not in candidates_by_no or dist < candidates_by_no[key]["distance_km"]:
                    candidates_by_no[key] = cand

    dedup_candidates = sorted(candidates_by_no.values(), key=lambda x: x["distance_km"])
    return {
        "matched": matched, "unmatched": unmatched,
        "good": good, "candidates": dedup_candidates,
    }


def _process_eksklyuziv(xlsx_bytes: bytes, radius_km: float):
    """
    HAMMA blokловchi ish (tarmoq + fayl parsing) shu yerda — run_read() orqali
    alohida oqimda ishlaydi, botning asosiy event loop'ini to'xtatib qo'ymaydi.

    MUHIM (2026-08): bir firmaning BIR NECHTA dorisi bo'lsa, xlsx faylda
    har bir dori ALOHIDA varaqda keladi. Endi BARCHA varaqlar o'qiladi va
    HAR BIR dori UCHUN alohida tahlil qilinadi (bir dorining "yaxshi
    ketmoqda" filiali boshqa doriga aralashtirilmaydi). Lokatsiya jadvali
    (Sheet1) esa faqat BIR MARTA yuklanadi va barcha dorilar uchun qayta
    ishlatiladi — tezlik va tarmoq yuklamasi uchun.

    Qaytaradi: dict(products=[{product_name, matched, unmatched, good,
    candidates}, ...], xlsx_bytes) — "products" ro'yxatida bitta yoki
    bir nechta element bo'lishi mumkin (faylda nechta dori bo'lsa shuncha).
    """
    products_raw = _parse_turnover_workbook(xlsx_bytes)
    if not products_raw:
        raise ValueError("Faylda hech qanday tahlil qilinadigan ma'lumot topilmadi.")

    loc_df = _load_location_df()
    name_choices = loc_df["Nomi (UZ)"].astype(str).tolist()
    name_to_row = {}
    for _, row in loc_df.iterrows():
        nom = _norm_name(row["Nomi (UZ)"])
        if nom and nom not in name_to_row:
            name_to_row[nom] = row

    products = []
    for product_name, turnover_rows in products_raw:
        analysis = _analyze_product(turnover_rows, loc_df, name_choices, name_to_row, radius_km)
        products.append({"product_name": product_name, **analysis})

    xlsx_out = _build_report_xlsx(products, radius_km)

    return {
        "products": products,
        "xlsx_bytes": xlsx_out,
    }


def _build_report_xlsx(products, radius_km):
    """
    MUHIM (2026-08): endi BITTA dori o'rniga `products` — bir yoki bir
    nechta dori natijalari ro'yxatini qabul qiladi (har biri {product_name,
    matched, unmatched, good, candidates}). Har ikki varaqqa ("Filiallar
    holati" va "Yaqin nomzod aptekalar") birinchi ustun sifatida "Dori"
    qo'shildi — shunda bir nechta dori bo'lsa ham, qaysi qator qaysi
    doriga tegishli ekani aniq ko'rinadi, va har bir dori o'z rangida
    (toifa bo'yicha) alohida ajratiladi. Bitta dori bo'lsa ham xatosiz
    ishlaydi — "Dori" ustuni shunchaki bitta qiymatga ega bo'ladi.
    """
    FONT = "Arial"
    HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
    HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=11)
    TITLE_FONT = Font(name=FONT, bold=True, size=14)
    NORMAL_FONT = Font(name=FONT, size=10)
    BOLD_FONT = Font(name=FONT, bold=True, size=10)
    THIN = Side(style="thin", color="D0D0D0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CAT_COLORS = {
        "Yaxshi ketmoqda": "C6EFCE",
        "Ketmagan (zaxira bor, sotuv yo'q)": "FFC7CE",
        "Sekin ketmoqda": "FFEB9C",
        "Yangi berilgan (hali erta baholash uchun)": "D9D9D9",
    }

    wb = openpyxl.Workbook()
    s2 = wb.active
    s2.title = "Filiallar holati"
    headers2 = ["Dori", "Filial", "Davr boshi qoldiq", "Sotildi (dona)", "Ko'chirildi",
                "Davr oxiri qoldiq", "Toifa", "Lokatsiya (Sheet1)", "Latitude", "Longitude"]
    for i, h in enumerate(headers2, 1):
        c = s2.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    s2.freeze_panes = "A2"
    row_i = 2
    for p in products:
        product_name = p["product_name"]
        for t in sorted(p["matched"], key=lambda x: -abs(x["sotuv"])):
            vals = [product_name, t["filial_raw"], t["ost_boshi"], abs(t["sotuv"]), t["peremeshenie"],
                    t["ost_oxiri"], t["category"], t["loc_nom"], t["lat"], t["lon"]]
            for i, v in enumerate(vals, 1):
                c = s2.cell(row=row_i, column=i, value=v)
                c.font = BOLD_FONT if i == 1 else NORMAL_FONT
                c.border = BORDER
                c.fill = PatternFill("solid", fgColor=CAT_COLORS.get(t["category"], "FFFFFF"))
            row_i += 1
        for t in p["unmatched"]:
            vals = [product_name, t["filial_raw"], t["ost_boshi"], abs(t["sotuv"]), t["peremeshenie"], t["ost_oxiri"],
                    f"Lokatsiya aniq topilmadi (eng yaqin taxmin: {t.get('best_guess','')}, {t.get('match_score',0)}%)",
                    "", "", ""]
            for i, v in enumerate(vals, 1):
                c = s2.cell(row=row_i, column=i, value=v)
                c.font = BOLD_FONT if i == 1 else NORMAL_FONT
                c.border = BORDER
                c.fill = PatternFill("solid", fgColor="BFBFBF")
            row_i += 1
    widths2 = [22, 28, 16, 14, 12, 16, 40, 34, 12, 12]
    for i, w in enumerate(widths2, 1):
        s2.column_dimensions[get_column_letter(i)].width = w

    s3 = wb.create_sheet("Yaqin nomzod aptekalar")
    title = (f"Har bir dori HALI YO'Q, lekin yaxshi ketayotgan filialga {radius_km} km "
             f"ichida joylashgan aptekalar" if len(products) > 1 else
             f"«{products[0]['product_name']}» hali YO'Q, lekin yaxshi ketayotgan filialga "
             f"{radius_km} km ichida joylashgan aptekalar")
    s3["A1"] = title
    s3["A1"].font = TITLE_FONT
    s3.merge_cells("A1:G1")
    headers3 = ["Dori", "Nomzod apteka (dori yo'q)", "Masofa (km)", "Eng yaqin \"yaxshi ketayotgan\" filial",
                "O'sha filialda sotilgan (dona)", "Nomzod Latitude", "Nomzod Longitude"]
    for i, h in enumerate(headers3, 1):
        c = s3.cell(row=3, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(wrap_text=True, vertical="center")
    s3.freeze_panes = "A4"
    row_i = 4
    for p in products:
        product_name = p["product_name"]
        for c in p["candidates"]:
            vals = [product_name, c["candidate_nom"], c["distance_km"], c["good_filial"], c["good_sold"],
                    c["candidate_lat"], c["candidate_lon"]]
            for i, v in enumerate(vals, 1):
                cell = s3.cell(row=row_i, column=i, value=v)
                cell.font = BOLD_FONT if i == 1 else NORMAL_FONT
                cell.border = BORDER
            row_i += 1
    widths3 = [22, 34, 12, 34, 20, 14, 14]
    for i, w in enumerate(widths3, 1):
        s3.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─── Telegram handlerlar ───────────────────────────────────────────────────

async def cmd_eksklyuziv(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """/eksklyuziv [radius_km] — admin buyrug'i: turnover xlsx faylni so'raydi."""
    if update.effective_user.id not in ADMIN_IDS:
        await update.message.reply_text("❌ Ruxsat yo'q.")
        return None

    radius = DEFAULT_RADIUS_KM
    if ctx.args:
        try:
            radius = float(ctx.args[0].replace(",", "."))
        except (ValueError, IndexError):
            pass
    ctx.user_data["eks_radius"] = radius

    await update.message.reply_text(
        f"📦 *Eksklyuziv dori tahlili*\n\n"
        f"Ushbu dorining aylanma (turnover) hisobotini *.xlsx* fayl sifatida yuboring "
        f"(\"Оборот по сети\" formatida — filial, qoldiq, sotuv ustunlari bilan).\n\n"
        f"📍 Radius: *{radius} km* (o'zgartirish uchun: `/eksklyuziv 1.5`)",
        parse_mode="Markdown",
    )
    return EKS_WAIT_XLSX


async def eks_receive_xlsx(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Turnover xlsx faylni qabul qiladi, tahlil qiladi va natijani qaytaradi."""
    if update.effective_user.id not in ADMIN_IDS:
        return None

    if not update.message.document or not update.message.document.file_name.lower().endswith(".xlsx"):
        await update.message.reply_text("❌ Iltimos, .xlsx fayl yuboring.")
        return EKS_WAIT_XLSX

    radius = ctx.user_data.get("eks_radius", DEFAULT_RADIUS_KM)
    msg = await update.message.reply_text("⏳ Tahlil qilinmoqda (bir necha soniya)...")

    try:
        file = await update.message.document.get_file()
        xlsx_bytes = bytes(await file.download_as_bytearray())
        result = await run_read(_process_eksklyuziv, xlsx_bytes, radius)
    except Exception as e:
        logger.exception("[EKS] Tahlil xatosi")
        await msg.edit_text(f"❌ Xatolik yuz berdi: {e}")
        return None

    products = result["products"]

    # MUHIM (2026-08): faylda BIR NECHTA dori (varaq) bo'lsa, har biri
    # uchun ALOHIDA qisqa xulosa ko'rsatiladi (bitta dori bo'lsa — avvalgi
    # kabi, bitta bloк chiqadi). Fayl esa doim BITTA, barcha dorilarni
    # o'z ichiga olgan xlsx sifatida yuboriladi (_build_report_xlsx orqali).
    if len(products) == 1:
        p = products[0]
        n_matched, n_unmatched = len(p["matched"]), len(p["unmatched"])
        n_good, n_cand = len(p["good"]), len(p["candidates"])
        lines = [
            f"✅ *«{p['product_name']}»* tahlili tayyor",
            f"",
            f"🏪 Jami filial (dori bor): {n_matched + n_unmatched}",
            f"📍 Lokatsiyasi topildi: {n_matched}" + (f" (⚠️ {n_unmatched} tasi topilmadi)" if n_unmatched else ""),
            f"🟢 Yaxshi ketmoqda: {n_good}",
            f"",
            f"🎯 {radius} km radiusda topilgan, dori HALI YO'Q nomzod aptekalar: *{n_cand}*",
        ]
    else:
        lines = [f"✅ *{len(products)} ta dori* tahlili tayyor (fayldagi {len(products)} ta varaq bo'yicha)", ""]
        total_cand = 0
        for p in products:
            n_matched, n_unmatched = len(p["matched"]), len(p["unmatched"])
            n_good, n_cand = len(p["good"]), len(p["candidates"])
            total_cand += n_cand
            lines.append(
                f"📦 *«{p['product_name']}»*: {n_matched + n_unmatched} filial"
                + (f" (⚠️ {n_unmatched} lokatsiyasi topilmadi)" if n_unmatched else "")
                + f", 🟢 yaxshi ketmoqda: {n_good}, 🎯 nomzod aptekalar: {n_cand}"
            )
        lines.append("")
        lines.append(f"📍 Radius: {radius} km · Jami nomzod aptekalar (barcha dorilar bo'yicha): *{total_cand}*")
    await msg.edit_text("\n".join(lines), parse_mode="Markdown")

    first_name = products[0]["product_name"][:20]
    suffix = f"_va_{len(products)-1}_boshqa" if len(products) > 1 else ""
    out_name = f"eksklyuziv_{first_name}{suffix}.xlsx".replace(" ", "_")
    file_io = io.BytesIO(result["xlsx_bytes"])
    file_io.name = out_name
    await ctx.bot.send_document(chat_id=update.effective_chat.id, document=file_io, filename=out_name)

    return None


def get_eks_states():
    return {
        EKS_WAIT_XLSX: [
            MessageHandler(filters.Document.ALL, eks_receive_xlsx),
        ],
    }
